# pylint: disable=C0114
from openpyxl import load_workbook
from config import user, target_mouth, location
import pandas as pd

# 讀取 Excel 文件
#user = "08020020"
#target_mouth = "04"
file_path = rf"D:\LongTermCare\WatchData\{location}\{user}\{user}_{target_mouth}month.xlsx" 
excel_data = pd.ExcelFile(file_path)

# 初始化一個字典來儲存結果
sheet_results = {}

# 對每個 sheet 執行計算
for sheet_name in excel_data.sheet_names:
    # 只處理包含"30mins"的工作表
    if "30mins" in sheet_name: 
        # 讀取每個sheet的數據
        sheet_data = excel_data.parse(sheet_name)

        if sheet_data.empty:
            print(f'Sheet "{sheet_name}" is empty, pass.')
            continue

        # 選擇 Completion rate(%) > 50 的行數
        above_50 = sheet_data[sheet_data['Completion rate(%)'] > 50]

        # 計算百分比，避免除以0
        percentage = (len(above_50) / len(sheet_data)) * 100 if len(sheet_data) > 0 else 0

        # 存儲結果
        sheet_results[sheet_name] = percentage



# 輸出每個 sheet 的結果
print("\n ===計算結果=== ")
for sheet, percent in sheet_results.items():
    print(f"Sheet '{sheet}' 的百分比 (Completion rate > 50): {percent:.3f}%")

# === 寫入原先的excel檔案 ===
# 將結果轉換為 Dataframe
result_df = pd.DataFrame(sheet_results.items(), columns=['Date', 'Completion rate > 50(%)'])
result_df['Date'] = result_df['Date'].str.replace('_30mins', '', regex=False)
print("\n 轉換後的結果：")
print(result_df)

# 將結果寫入新的 Excel 文件
with pd.ExcelWriter(file_path, if_sheet_exists='replace', mode='a') as writer:  # mode='w' 會覆蓋原檔案
    result_df.to_excel(writer, sheet_name='Summary', index=False) 

# 使用Openpyxl 調整sheet位置
workbook = load_workbook(file_path)
# 先確保 'Summary' 工作表存在
if 'Summary' in workbook.sheetnames:
    summary_sheet = workbook['Summary']
    # 將 'Summary' 工作表移到第一個位置
    workbook.move_sheet(summary_sheet, offset=-len(workbook.sheetnames) + 1)
    # 保存檔案
    workbook.save(file_path)
    print("工作表 'Summary' 已移到第一個位置！")
    print("結果儲存在:", file_path)
else:
    print("工作表 'Summary' 不存在！")

