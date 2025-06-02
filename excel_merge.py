# pylint: disable=C0114
from openpyxl import load_workbook
from config import user, target_mouth, location
import pandas as pd

# 讀取 Excel 文件
file_path = rf"D:\LongTermCare\WatchData\{location}\{user}\file\\total_count\Summary\{user}_total_count.xlsx"

# 讀取所有 sheet
all_sheets = pd.read_excel(file_path, sheet_name=None)

# 將所有 sheet 的資料合併
merged_df = pd.concat(
    [df[['日期', '資料筆數']] for df in all_sheets.values()],
    ignore_index=True
)

print(merged_df)

# === 寫入原先的excel檔案 ===

# 將結果寫入新的 Excel 文件
with pd.ExcelWriter(file_path, if_sheet_exists='replace', mode='a') as writer:  # mode='w' 會覆蓋原檔案
    merged_df.to_excel(writer, sheet_name='Summary', index=False) 

# 使用Openpyxl 調整sheet位置
workbook = load_workbook(file_path)
# 先確保 'Summary' 工作表存在
if 'Summary' in workbook.sheetnames:
    summary_sheet = workbook['Summary']
    # 將 'Summary' 工作表移到第一個位置
    workbook.move_sheet(summary_sheet, offset=-len(workbook.sheetnames) + 1)
    # 保存檔案
    workbook.save(file_path)
    print("工作表 'Summary' 已移到第一個位置！")
    print("結果儲存在:", file_path)
else:
    print("工作表 'Summary' 不存在！")
